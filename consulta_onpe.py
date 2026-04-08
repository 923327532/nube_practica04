"""
CONSULTA MASIVA ONPE - Elecciones Generales 2026
Lee datos_onpe.xlsx, consulta la web ONPE con Chrome real y rellena el Excel.

Instalacion local:
    pip install selenium openpyxl

Uso local:
    python consulta_onpe.py

En Docker:
    docker build -t onpe .
    docker run -v "%cd%":/app onpe
"""

import time
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException

ARCHIVO_EXCEL = "onpe.xlsx"
URL_ONPE      = "https://consultaelectoral.onpe.gob.pe/inicio"

COLOR_MIEMBRO   = "FFF3CD"
COLOR_CIUDADANO = "E8F5E9"
COLOR_ERROR     = "FFCCCC"

# ─────────────────────────────────────────────────────────
#  CHROME DRIVER
# ─────────────────────────────────────────────────────────
def crear_driver(headless=False):
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--disable-infobars")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36")
    driver = webdriver.Chrome(options=opts)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    })
    return driver


# ─────────────────────────────────────────────────────────
#  INTERCEPTAR RESPUESTA JSON via CDP Network
# ─────────────────────────────────────────────────────────
def activar_interceptor(driver):
    """Inyecta JS para capturar la respuesta de /consulta/definitiva"""
    driver.execute_script("""
        window._onpe_data = null;
        window._onpe_error = null;
        const _orig = window.fetch;
        window.fetch = async function(...args) {
            const url = (args[0] || '').toString();
            const resp = await _orig.apply(this, args);
            if (url.includes('consulta/definitiva') || url.includes('busqueda/dni')) {
                try {
                    const clone = resp.clone();
                    const json = await clone.json();
                    if (json && json.success && json.data && json.data.nombres) {
                        window._onpe_data = json.data;
                    } else if (json && json.data && json.data.token) {
                        window._onpe_token = json.data.token;
                    }
                } catch(e) {}
            }
            return resp;
        };
    """)


# ─────────────────────────────────────────────────────────
#  CONSULTAR UN DNI
# ─────────────────────────────────────────────────────────
def consultar_dni(driver, dni, intento=1):
    try:
        driver.get(URL_ONPE)

        # Esperar que la pagina cargue (Angular necesita tiempo)
        time.sleep(3)

        # Verificar si hay error 500
        if "Error interno" in driver.page_source or "500" in driver.title:
            print(f" [servidor ONPE caído, reintentando en 15s]", end="", flush=True)
            time.sleep(15)
            if intento < 3:
                return consultar_dni(driver, dni, intento + 1)
            return None

        # Activar interceptor fetch
        activar_interceptor(driver)
        driver.execute_script("window._onpe_data = null;")

        # Buscar campo input (Angular genera inputs dinamicamente)
        wait = WebDriverWait(driver, 20)
        
        # Intentar varios selectores para el input del DNI
        input_el = None
        for selector in [
            "input[maxlength='8']",
            "input[type='text']",
            "input[formcontrolname='numeroDocumento']",
            "input[placeholder*='DNI']",
            "input[placeholder*='dni']",
            "input[name*='dni']",
            "input[name*='documento']",
            "input",
        ]:
            try:
                input_el = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
                if input_el:
                    break
            except TimeoutException:
                continue

        if not input_el:
            print(f" [no se encontró campo DNI]", end="")
            return None

        # Limpiar e ingresar DNI
        input_el.clear()
        time.sleep(0.5)
        input_el.send_keys(dni)
        time.sleep(1)

        # Esperar que recaptcha v3 se inicialice (invisible, automatico)
        time.sleep(2)

        # Buscar boton CONSULTAR
        boton = None
        for selector in [
            "button[type='submit']",
            "button.btn-primary",
            "button.consultar",
        ]:
            try:
                boton = driver.find_element(By.CSS_SELECTOR, selector)
                if boton:
                    break
            except:
                continue

        # Si no encontramos por CSS, buscar por texto
        if not boton:
            botones = driver.find_elements(By.TAG_NAME, "button")
            for b in botones:
                texto = (b.text or "").upper().strip()
                if "CONSULTAR" in texto or "BUSCAR" in texto or "CONSULTA" in texto:
                    boton = b
                    break

        if not boton:
            print(f" [no se encontró botón CONSULTAR]", end="")
            # Debug: mostrar todos los botones
            botones = driver.find_elements(By.TAG_NAME, "button")
            print(f" [botones disponibles: {[b.text for b in botones]}]", end="")
            return None

        # Click en CONSULTAR
        driver.execute_script("arguments[0].click();", boton)
        print(f" [click consultar]", end="", flush=True)

        # Esperar respuesta (hasta 25 segundos para recaptcha + respuesta)
        for i in range(25):
            time.sleep(1)
            data = driver.execute_script("return window._onpe_data;")
            if data and isinstance(data, dict) and data.get("nombres"):
                return data
            # Ver si hubo error visible en pantalla
            if "no encontrado" in driver.page_source.lower() or \
               "dni incorrecto" in driver.page_source.lower():
                return None

        print(f" [timeout esperando respuesta]", end="")
        return None

    except Exception as e:
        print(f" [excepcion: {type(e).__name__}: {str(e)[:60]}]", end="")
        return None


# ─────────────────────────────────────────────────────────
#  EXCEL
# ─────────────────────────────────────────────────────────
def detectar_columnas(ws):
    mapa = {}
    for col in range(1, 20):
        val = ws.cell(row=1, column=col).value
        if val:
            mapa[str(val).strip().lower()] = col
    return mapa


def escribir_fila(ws, row, cols, info, error):
    if error or info is None:
        bg = COLOR_ERROR
        nombres_val = "ERROR / No encontrado"
        miembro_val = "ERROR"
        ubigeo_val  = ""
        dir_val     = ""
    else:
        es_m = info.get("miembroMesa", False)
        bg   = COLOR_MIEMBRO if es_m else COLOR_CIUDADANO
        nombres_val = f"{info.get('nombres','').strip()} {info.get('apellidos','').strip()}".strip()
        miembro_val = "SI" if es_m else "NO"
        ubigeo_val  = info.get("ubigeo", "")
        partes = [p for p in [
            info.get("localVotacion", ""),
            info.get("direccion", ""),
            info.get("referencia", ""),
        ] if p]
        dir_val = " | ".join(partes)

    fill = PatternFill("solid", start_color=bg)
    f    = Font(size=10)
    izq  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    cen  = Alignment(horizontal="center", vertical="center")

    ws.cell(row=row, column=cols.get("dni", 1)).fill = fill

    for k in ["nombres", "nombre completo", "nombre"]:
        if k in cols:
            c = ws.cell(row=row, column=cols[k], value=nombres_val)
            c.fill = fill; c.font = f; c.alignment = izq; break

    for k in ["mienbro de mesa", "miembro de mesa"]:
        if k in cols:
            c = ws.cell(row=row, column=cols[k], value=miembro_val)
            c.fill = fill; c.font = f; c.alignment = cen; break

    for k in ["ubicación", "ubicacion", "ubigeo"]:
        if k in cols:
            c = ws.cell(row=row, column=cols[k], value=ubigeo_val)
            c.fill = fill; c.font = f; c.alignment = izq; break

    for k in ["direccion", "dirección"]:
        if k in cols:
            c = ws.cell(row=row, column=cols[k], value=dir_val)
            c.fill = fill; c.font = f; c.alignment = izq; break


# ─────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  CONSULTA MASIVA ONPE - Elecciones Generales 2026")
    print("=" * 55)

    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb.active
    cols = detectar_columnas(ws)
    print(f"\nColumnas: {list(cols.keys())}")

    if "dni" not in cols:
        print("ERROR: No se encontró columna DNI"); sys.exit(1)

    filas = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=cols["dni"]).value
        if val:
            filas.append((row, str(val).strip().zfill(8)))

    print(f"DNIs a procesar: {len(filas)}")
    print("Abriendo Chrome... (no lo cierres)\n" + "-" * 55)

    # En Docker usamos headless, localmente con ventana
    import os
    headless = os.environ.get("HEADLESS", "0") == "1"
    driver = crear_driver(headless=headless)

    ok = errores = 0

    try:
        for idx, (row, dni) in enumerate(filas, 1):
            print(f"[{idx}/{len(filas)}] DNI: {dni}", end="", flush=True)
            ws.row_dimensions[row].height = 22

            info = consultar_dni(driver, dni)

            if info and info.get("nombres"):
                es_m   = info.get("miembroMesa", False)
                nombre = f"{info.get('nombres','').strip()} {info.get('apellidos','').strip()}"
                print(f" → {nombre} | {'MIEMBRO ⭐' if es_m else 'No miembro'}")
                escribir_fila(ws, row, cols, info, error=False)
                ok += 1
            else:
                print(" → No encontrado")
                escribir_fila(ws, row, cols, None, error=True)
                errores += 1

            wb.save(ARCHIVO_EXCEL)
            time.sleep(2)

    finally:
        driver.quit()

    anchos = {"dni":14,"nombres":32,"mienbro de mesa":16,"miembro de mesa":16,
              "ubicación":34,"ubicacion":34,"direccion":42}
    for n, ci in cols.items():
        ws.column_dimensions[get_column_letter(ci)].width = anchos.get(n, 18)
    wb.save(ARCHIVO_EXCEL)

    print(f"\n{'='*55}")
    print(f"LISTO. Exitosos: {ok} | Errores: {errores}")
    print(f"Archivo guardado: {ARCHIVO_EXCEL}")

if __name__ == "__main__":
    main()